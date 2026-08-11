from pathlib import Path
from shutil import copy2

import joblib
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn.ensemble import RandomForestClassifier


项目目录 = Path(__file__).resolve().parents[1]

原始数据目录 = 项目目录 / "data" / "raw"
模型输出目录 = 项目目录 / "outputs" / "models"
表格输出目录 = 项目目录 / "outputs" / "tables"
提交输出目录 = 项目目录 / "outputs" / "submission"

附件一文件 = 原始数据目录 / "附件一（训练集）.xlsx"
附件二文件 = 原始数据目录 / "附件二（测试集）.xlsx"
附件四原始文件 = 原始数据目录 / "附件四（Excel表）.xlsx"

附件四输出文件 = 提交输出目录 / "附件四（Excel表）.xlsx"
预测结果文件 = 表格输出目录 / "附件二_励磁波形分类结果.csv"
模型文件 = 模型输出目录 / "随机森林波形分类模型.joblib"

材料名称 = ["材料1", "材料2", "材料3", "材料4"]
波形名称 = ["正弦波", "三角波", "梯形波"]

类别编码 = {
    "正弦波": 1,
    "三角波": 2,
    "梯形波": 3
}

分位点 = np.linspace(0, 1, 41)


def 检查文件():
    文件列表 = [
        附件一文件,
        附件二文件,
        附件四原始文件
    ]

    for 文件 in 文件列表:
        if not 文件.exists():
            raise FileNotFoundError(
                f"找不到文件：{文件}"
            )


def 提取一阶差分特征(磁通密度):
    磁通密度 = np.asarray(
        磁通密度,
        dtype=float
    )

    if 磁通密度.ndim != 2:
        raise ValueError(
            "磁通密度数据必须为二维数组"
        )

    if 磁通密度.shape[1] != 1024:
        raise ValueError(
            f"磁通密度采样点数为"
            f"{磁通密度.shape[1]}，不是1024"
        )

    if not np.isfinite(磁通密度).all():
        raise ValueError(
            "磁通密度中存在缺失值或无穷值"
        )

    最小值 = 磁通密度.min(
        axis=1,
        keepdims=True
    )

    最大值 = 磁通密度.max(
        axis=1,
        keepdims=True
    )

    峰峰值 = 最大值 - 最小值

    恒定波形位置 = np.where(
        峰峰值.ravel() <= 1e-12
    )[0]

    if len(恒定波形位置) > 0:
        raise ValueError(
            f"发现{len(恒定波形位置)}条"
            f"峰峰值接近0的恒定波形"
        )

    归一化波形 = (
        (磁通密度 - 最小值) / 峰峰值
    )

    一阶差分 = np.diff(
        归一化波形,
        axis=1,
        append=归一化波形[:, :1]
    )

    一阶差分 = 一阶差分 * 1024

    差分分位数 = np.quantile(
        一阶差分,
        分位点,
        axis=1
    ).T

    差分均值 = 一阶差分.mean(
        axis=1,
        keepdims=True
    )

    差分标准差 = 一阶差分.std(
        axis=1,
        keepdims=True
    )

    中心差分 = 一阶差分 - 差分均值
    安全标准差 = 差分标准差 + 1e-12

    差分偏度 = np.mean(
        (中心差分 / 安全标准差) ** 3,
        axis=1,
        keepdims=True
    )

    差分峰度 = (
        np.mean(
            (中心差分 / 安全标准差) ** 4,
            axis=1,
            keepdims=True
        )
        - 3
    )

    差分最小值 = 一阶差分.min(
        axis=1,
        keepdims=True
    )

    差分最大值 = 一阶差分.max(
        axis=1,
        keepdims=True
    )

    特征矩阵 = np.hstack([
        差分分位数,
        差分标准差,
        差分偏度,
        差分峰度,
        差分最小值,
        差分最大值
    ])

    分位数列名 = [
        f"差分分位数_{int(数值 * 100):03d}"
        for 数值 in 分位点
    ]

    特征列名 = 分位数列名 + [
        "差分标准差",
        "差分偏度",
        "差分峰度",
        "差分最小值",
        "差分最大值"
    ]

    return pd.DataFrame(
        特征矩阵,
        columns=特征列名
    )


def 读取附件一特征():
    特征列表 = []
    标签列表 = []

    for 当前材料 in 材料名称:
        print(f"正在读取附件一：{当前材料}")

        当前数据 = pd.read_excel(
            附件一文件,
            sheet_name=当前材料,
            engine="openpyxl"
        )

        当前标签 = 当前数据.iloc[
            :,
            3
        ].astype(str)

        非法标签 = set(
            当前标签.unique()
        ) - set(波形名称)

        if 非法标签:
            raise ValueError(
                f"{当前材料}存在非法标签："
                f"{非法标签}"
            )

        当前磁通密度 = 当前数据.iloc[
            :,
            4:
        ].to_numpy(dtype=float)

        当前特征 = 提取一阶差分特征(
            当前磁通密度
        )

        特征列表.append(当前特征)
        标签列表.append(当前标签)

    训练特征 = pd.concat(
        特征列表,
        ignore_index=True
    )

    训练标签 = pd.concat(
        标签列表,
        ignore_index=True
    )

    return 训练特征, 训练标签


def 读取附件二特征():
    print("正在读取附件二")

    附件二数据 = pd.read_excel(
        附件二文件,
        sheet_name="测试集",
        engine="openpyxl"
    )

    样本序号 = 附件二数据.iloc[
        :,
        0
    ].astype(int)

    磁通密度 = 附件二数据.iloc[
        :,
        4:
    ].to_numpy(dtype=float)

    测试特征 = 提取一阶差分特征(
        磁通密度
    )

    return 样本序号, 测试特征


def 训练最终模型(训练特征, 训练标签):
    print("正在使用全部附件一数据训练最终模型")

    模型 = RandomForestClassifier(
        n_estimators=300,
        max_features="sqrt",
        class_weight="balanced",
        random_state=42,
        n_jobs=-1
    )

    模型.fit(
        训练特征,
        训练标签
    )

    return 模型


def 生成预测结果(模型, 样本序号, 测试特征):
    预测波形 = 模型.predict(
        测试特征
    )

    预测结果 = pd.DataFrame({
        "序号": 样本序号,
        "预测波形": 预测波形
    })

    预测结果["分类数字"] = (
        预测结果["预测波形"]
        .map(类别编码)
        .astype(int)
    )

    return 预测结果


def 输出统计信息(预测结果):
    print("\n附件二预测完成")
    print("\n三种波形数量：")

    波形统计 = (
        预测结果["预测波形"]
        .value_counts()
        .reindex(
            波形名称,
            fill_value=0
        )
    )

    print(波形统计)

    指定序号 = [
        1,
        5,
        15,
        25,
        35,
        45,
        55,
        65,
        75,
        80
    ]

    指定结果 = 预测结果[
        预测结果["序号"].isin(
            指定序号
        )
    ].sort_values("序号")

    print("\n题目指定样本的分类结果：")
    print(
        指定结果.to_string(
            index=False
        )
    )


def 保存结果(模型, 预测结果):
    模型输出目录.mkdir(
        parents=True,
        exist_ok=True
    )

    表格输出目录.mkdir(
        parents=True,
        exist_ok=True
    )

    提交输出目录.mkdir(
        parents=True,
        exist_ok=True
    )

    joblib.dump(
        模型,
        模型文件
    )

    预测结果.to_csv(
        预测结果文件,
        index=False,
        encoding="utf-8-sig"
    )

    copy2(
        附件四原始文件,
        附件四输出文件
    )

    工作簿 = load_workbook(
        附件四输出文件
    )

    工作表 = 工作簿["Sheet1"]

    for _, 结果行 in 预测结果.iterrows():
        Excel行号 = int(
            结果行["序号"]
        ) + 1

        工作表.cell(
            row=Excel行号,
            column=2,
            value=int(
                结果行["分类数字"]
            )
        )

    工作簿.save(
        附件四输出文件
    )

    print("\n结果已经保存：")
    print(f"模型：{模型文件}")
    print(f"预测表：{预测结果文件}")
    print(f"附件四：{附件四输出文件}")


def 主程序():
    检查文件()

    训练特征, 训练标签 = (
        读取附件一特征()
    )

    print(
        f"训练样本数：{len(训练标签)}"
    )

    print(
        f"特征数量：{训练特征.shape[1]}"
    )

    最终模型 = 训练最终模型(
        训练特征,
        训练标签
    )

    样本序号, 测试特征 = (
        读取附件二特征()
    )

    预测结果 = 生成预测结果(
        最终模型,
        样本序号,
        测试特征
    )

    输出统计信息(
        预测结果
    )

    保存结果(
        最终模型,
        预测结果
    )


if __name__ == "__main__":
    主程序()